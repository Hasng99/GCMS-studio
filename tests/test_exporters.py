import io

import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

from src.exporters import _apply_quality_flags, _safe_sheet_name, selected_hits_to_xlsx_bytes


def test_safe_sheet_name_sanitizes_invalid_characters_and_dedupes() -> None:
    used: set[str] = set()
    first = _safe_sheet_name("sample: a/b", used)
    second = _safe_sheet_name("sample: a/b", used)

    assert first == "sample_ a_b"
    assert second != first
    assert first in used and second in used


def test_selected_hits_to_xlsx_bytes_splits_sheets_and_drops_sample_name() -> None:
    results_by_sample = {
        "sample A": pd.DataFrame({
            "sample_name": ["sample A"],
            "canonical_name": ["Hexanal"],
            "area": [100],
        }),
        "sample B": pd.DataFrame({
            "sample_name": ["sample B"],
            "canonical_name": ["Octanal"],
            "area": [200],
        }),
    }

    workbook = openpyxl.load_workbook(io.BytesIO(selected_hits_to_xlsx_bytes(results_by_sample)))

    assert workbook.sheetnames == ["sample A", "sample B"]
    header = [cell.value for cell in next(workbook["sample A"].iter_rows(min_row=1, max_row=1))]
    assert header == ["canonical_name", "area"]


def test_apply_quality_flags_colors_only_flagged_sample_area_cell() -> None:
    frame = pd.DataFrame({
        "canonical_name": ["Hexanal"],
        "area::rep-1": [100],
        "area::rep-2": [110],
        "quality_flag_samples": ["rep-2"],
    })
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in dataframe_to_rows(frame, index=False, header=True):
        sheet.append(row)

    _apply_quality_flags(sheet, frame)

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    flagged_column = header.index("area::rep-2") + 1
    other_column = header.index("area::rep-1") + 1
    assert sheet.cell(row=2, column=flagged_column).fill.fgColor.rgb.endswith("FFCDD2")
    assert sheet.cell(row=2, column=other_column).fill.fill_type is None
