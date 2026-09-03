from __future__ import annotations

import io
import re
from typing import Mapping

import pandas as pd
from openpyxl.styles import PatternFill

from .multi_sample import AREA_PREFIX
from .pipeline import PipelineResult


_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_QUALITY_FLAG_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")


def dataframe_to_csv_bytes(frame: pd.DataFrame) -> bytes:
    return frame.to_csv(index=False).encode("utf-8-sig")


def _safe_sheet_name(name: object, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name).strip()) or "sample"
    base = cleaned[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def selected_hits_to_xlsx_bytes(results_by_sample: Mapping[str, pd.DataFrame]) -> bytes:
    """샘플별 시트로 분리하고, 시트 안에서는 반복되는 sample_name 열을 생략한다."""
    output = io.BytesIO()
    used_sheet_names: set[str] = set()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sample_name, frame in results_by_sample.items():
            sheet_name = _safe_sheet_name(sample_name, used_sheet_names)
            frame.drop(columns=["sample_name"], errors="ignore").to_excel(
                writer, sheet_name=sheet_name, index=False
            )
    return output.getvalue()


def _apply_quality_flags(worksheet, frame: pd.DataFrame) -> None:
    """quality_flag_samples에 기록된 샘플의 Area 칸을 빨간색으로 표시한다."""
    if "quality_flag_samples" not in frame.columns:
        return
    column_index = {name: index + 1 for index, name in enumerate(frame.columns)}
    for row_offset, flagged in enumerate(frame["quality_flag_samples"].fillna(""), start=2):
        for sample_name in str(flagged).split(";"):
            if not sample_name:
                continue
            column = column_index.get(f"{AREA_PREFIX}{sample_name}")
            if column:
                worksheet.cell(row=row_offset, column=column).fill = _QUALITY_FLAG_FILL


def results_to_xlsx_bytes(
    result: PipelineResult,
    standards: pd.DataFrame,
    profile: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.peak_summary.to_excel(writer, sheet_name="peak_summary", index=False)
        result.selected_hits.to_excel(writer, sheet_name="selected_hits", index=False)
        result.rejected_hits.to_excel(writer, sheet_name="rejected_hits", index=False)
        result.all_hits.to_excel(writer, sheet_name="all_hits", index=False)
        standards.to_excel(writer, sheet_name="standards", index=False)
        profile.to_excel(writer, sheet_name="profile", index=False)
    return output.getvalue()


def multi_results_to_xlsx_bytes(
    results: Mapping[str, PipelineResult],
    standards: pd.DataFrame,
    profile: pd.DataFrame,
    *,
    comparison: pd.DataFrame | None = None,
    comparison_sheet: str = "comparison",
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if comparison is not None:
            sheet_name = comparison_sheet[:31]
            comparison.to_excel(writer, sheet_name=sheet_name, index=False)
            _apply_quality_flags(writer.sheets[sheet_name], comparison)
        pd.DataFrame({
            "sample_index": range(1, len(results) + 1),
            "sample_name": list(results),
        }).to_excel(writer, sheet_name="samples", index=False)
        for index, result in enumerate(results.values(), start=1):
            prefix = f"sample_{index}"
            result.peak_summary.to_excel(
                writer, sheet_name=f"{prefix}_summary", index=False
            )
            result.selected_hits.to_excel(
                writer, sheet_name=f"{prefix}_selected", index=False
            )
            result.rejected_hits.to_excel(
                writer, sheet_name=f"{prefix}_rejected", index=False
            )
            result.all_hits.to_excel(
                writer, sheet_name=f"{prefix}_all", index=False
            )
        standards.to_excel(writer, sheet_name="standards", index=False)
        profile.to_excel(writer, sheet_name="profile", index=False)
    return output.getvalue()
